import base64
import io
from openpyxl import load_workbook

from odoo import _, api, fields, models
from odoo.exceptions import UserError

class VendorProductImport(models.Model):
    _name = 'vendor.product.import'
    _inherit = ['mail.thread']

    name = fields.Char(string="Name", default="New", readonly="1")
    vendor_id = fields.Many2one('res.partner', string="Vendor", required=True)
    vendor_template = fields.Many2one('vendor.product.template', string="Vendor Template Formate", domain="[('vendor_id', '=', vendor_id)]")
    file_to_process = fields.Binary(string="File to Process")
    file_name = fields.Char(string="File Name")
    date = fields.Datetime(string="Date", default=fields.Datetime.now)
    state = fields.Selection(
        selection=[
            ('pending', "Pending"),
            ('processed', "Processed"),
            ('error', "Error")
        ],
        tracking=True,
        string="State",
        default='pending'
    )

    @api.onchange('vendor_id')
    def _onchange_vendor_id(self):
        if self.vendor_id:
            template = self.env['vendor.product.template'].search([('vendor_id', '=', self.vendor_id.id)], limit=1)
            self.vendor_template = template.id

    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            if vals.get('name', 'New') == 'New':
                vals['name'] = self.env['ir.sequence'].next_by_code('vendor.product.import') or 'New'

        return super().create(vals_list)

    # Acction methods

    def action_reset_to_pending(self):
        self.state = 'pending'

    def action_manual_process(self):
        """Process the uploaded Excel file to update or create products and log history."""
        if not self.file_to_process:
            raise UserError("Please upload an Excel file before processing.")

        if not self.vendor_template:
            raise UserError("No vendor template selected.")

        try:
            # Read the Excel file
            file_content = base64.b64decode(self.file_to_process)
            file_stream = io.BytesIO(file_content)
            workbook = load_workbook(file_stream, data_only=True)
            sheet = workbook.worksheets[0]

            # Get required field mappings from vendor template
            template_format_lines = self.vendor_template.template_formate_ids
            field_mapping = {line.file_header: line.odoo_field.name for line in template_format_lines}

            header_tuple = {header: idx for idx, header in enumerate(cell.value for cell in sheet[1])}
            missing_headers = set(field_mapping.keys()) - set(header_tuple.keys())
            if missing_headers:
                self.state = 'error'
                self.message_post(body=_(f"The uploaded file is missing required headers: {', '.join(missing_headers)}"))
                return

            products_to_create = []
            products_to_update = []
            import_references = []
            all_existing_products = self.env['product.template'].with_context(active_test=False).search([])
            all_existing_products = {product.product_unique_id : product for product in all_existing_products}

            for row in sheet.iter_rows(min_row=2, values_only=True):
                product_data = {field_mapping[header]: row[header_tuple.get(header)] for header in field_mapping}
                product_unique_id = product_data.get('product_unique_id')
                if not product_unique_id:
                    self.message_post(body=_("Missing 'product_unique_id' in the imported data."))
                    return

                existing_product = all_existing_products.get(product_unique_id)
                old_price = existing_product.list_price if existing_product else 0
                
                if existing_product:
                    existing_product.write(product_data)
                    products_to_update.append(existing_product.id)
                    new_price = existing_product.list_price
                else:
                    new_product = self.env['product.template'].create(product_data)
                    products_to_create.append(new_product.id)
                    new_price = new_product.list_price

                import_references.append({
                    'import_reference': f"{self.name}",
                    'date_of_import': fields.Date.today(),
                    'file_name': self.file_name,
                    'old_price': old_price,
                    'new_price': new_price,
                    'product_id': existing_product.id if existing_product else new_product.id
                })
            
            if import_references:
                self.env['product.import.histry'].create(import_references)

            self.message_post(body=_(f'Successfully processed {len(products_to_create)} new product(s) and updated {len(products_to_update)} existing product(s).'))
            self.state = 'processed'
            
        except Exception as e:
            self.state = 'error'
            self.message_post(body=_(f"An error occurred while processing the file: {str(e)}"))
